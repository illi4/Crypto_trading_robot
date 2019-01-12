import os
from openpyxl import load_workbook
import openpyxl

import argparse

def parce_folder():
    # Parse input params
    parser = argparse.ArgumentParser()
    parser.add_argument('--folder', type=str, help="Folder (path to results)")
    args, unknown = parser.parse_known_args()
    folder = getattr(args, 'folder')

    return folder


### Start
print('Use example: result_transform.py --folder="D:\Dropbox\Exchange\! ML new test results\btc"')

folder = parce_folder()
#cwd = os.getcwd()

column = 1

# Create the file 
filepath = "{}/results_processed.xlsx".format(folder)
wb = openpyxl.Workbook()
wb.save(filepath)

wb = load_workbook(filepath)
ws = wb.get_active_sheet()

for filename in os.listdir(folder):
    if filename.endswith(".log") and filename.find('summary') > 0:
        print (filename) 
        with open('%s\\%s' % (folder, filename), 'r+') as file:
            row = 1
            line_original = file.readline() 
            ws.cell(row = row, column = column).value = line_original
            row += 1
            while line_original !="":
                line_original = file.readline()
                
                # Write converted where possible 
                line = line_original.replace('\n', '').replace('\r', '')
                try: 
                    line = float(line)
                except: 
                    pass 
    
                line = ws.cell(row = row, column = column).value = line
                row += 1
        column += 1

wb.save(filepath)