################################ Libraries ############################################
## Standard libraries 

from time import localtime, strftime, sleep
import logging
import os
from openpyxl import load_workbook

### Logfile class 
class logfile(object):
    def __init__(self, market, type, codename = None, username = None):
        self.public = ['initialise', 'write', 'close', 'lprint']
        self.check_dirs(['logs_trade', 'system_msg'])       # check directories

        if username is None:
            username = ''

        trade_id = "".join([codename, ' ', username, ' ', market, ' ', strftime("%Y-%m-%d %H-%M-%S", localtime())])
        if type == 'buy' or type == 'trade': 
            self.filename = "logs_trade/" + trade_id + ".log"
            self.filename_summary = "logs_trade/" + trade_id + "_summary.log"
        else: 
            self.filename = "system_msg/" + trade_id + ".log"

        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.INFO)
        # Create a file handler
        self.handler = logging.FileHandler(self.filename)
        self.handler.setLevel(logging.INFO)
        # Create a logging format
        formatter = logging.Formatter('%(asctime)s: %(levelname)s: %(message)s')
        self.handler.setFormatter(formatter)
        # Add the handlers to the logger
        self.logger.addHandler(self.handler)

    def check_dirs(self, dirlist): 
        # Check if required directories exist and create them if not 
        for dir in dirlist:     
            directory = os.path.dirname(dir)
            try:
                os.stat(dir)
            except:
                os.mkdir(dir) 
    
    def write(self, msg):  
        self.logger.info(msg)

    def write_summary(self, results, balance_snapshot, start_time, trade_time, margin,
        codename = '', col_no = None, testfile = None, tabname = 'results', params_info = ''):

        ''' # For testing
        f = open(self.filename_summary, 'w')
        a = ''
        for elem in results:
            a += str(elem)
        f.write(str(results))
        f.write(str(balance))
        f.write('---')
        '''
        '''
        # Export multiple into excel  # does not work when too many threads 
        if col_no is not None and testfile is not None:
            wb = load_workbook(testfile)  # check how to add to specific column
            ws = wb[tabname]
            # Write in the excel   params_info
            ws.cell(row = 1, column = col_no).value = params_info
            ws.cell(row = 2, column = col_no).value = codename
            ws.cell(row = 3, column = col_no).value = 'End balance'
            ws.cell(row = 4, column = col_no).value = balance
            ws.cell(row = 5, column = col_no).value = 'Results (%, no margin):'
            for row, val in enumerate(results):
                ws.cell(row=row+7, column=col_no).value = val[2]
            wb.save(testfile)
        '''

        f = open(self.filename_summary, 'w')
        # Writing the trades info
        f.write(codename)
        f.write("\n\n")
        f.write(params_info)
        f.write('\nMargin: ')
        f.write(str(margin))
        # Writing the margin info
        str_period = "\nTested from {} to {}".format(start_time, trade_time)
        f.write(str_period)
        #f.write('\nTested up to: ')
        #f.write(trade_time)
        # Calc of total cumulative %
        total_result = 1
        for elem in results:
            total_result = total_result*(1 + margin*float(elem[2])/float(100))
            #f.write('total calc {}\n'.format(total_result))
            
        # Get the max drawdown 
        #max_drawdown = min(results, key=lambda x: x[2])
        sorted_results = sorted(results, key=lambda x: x[2])
        max_drawdowns = [item[2] for item in sorted_results] 
        max_drawdowns = ['{:.2f}'.format(x) for x in max_drawdowns][:3]

        balance_snapshot = balance_snapshot*total_result
        
        # Writing totals
        f.write("\ntotal x (on margin)\n")
        f.write("{}\n".format(total_result))
        f.write("\nMax drawdowns:\n")
        f.write("{}\n".format(max_drawdowns))
        f.write("\ncumulative balance (trade base currency, on margin)\n")
        f.write("{}\n".format(balance_snapshot))

        # Details on %
        f.write("\n\nReturned % (no margin)\n")
        for elem in results:
            f.write(str(elem[2]))
            f.write("\n")
        # Writing the dates and the results
        f.write('\nDetailed returns per dates:\n')
        for elem in results: 
            str_write = ", ".join([elem[0], elem[1], str(elem[2])])
            f.write(str_write)
            f.write("\n")
        f.close()

    '''
    def close_and_exit(self):
        self.logger.removeHandler(self.handler)
        del self.logger, self.handler
        exit(0)
    '''

    # Close handler
    def close_handler(self):
        self.logger.removeHandler(self.handler)
        del self.logger, self.handler

    # Log and print 
    def lprint(self, arr):
        msg = ' '.join([''+ str(x) for x in arr])
        try: 
            self.write(msg)
            print(msg)
        except: 
            print('Failed to write output due to the IO error')
            print(msg)
        
    def close(self):
        self.logger.removeHandler(self.handler)
        del self.logger, self.handler
