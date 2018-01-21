##################################################################################################################
# Use: python robot.py exchange simulation_flag(s/r/sns/rns) basic_currency altcoin entry_price TP% SL% [limit_of_amount_to sell] [sell_portion]
# s - simulation with stop-loss
# r - real mode with stop-loss 
# sns - simulation without stop loss and only with trailing stop on profit 
# rns - real mode without stop loss and only with trailing stop on profit 
#
# Example: trade 100 LTC (vs BTC) bought at 0.0017 with the target profit of 18% and stop-loss 5% in simulation mode
# > python robot.py btrx s BTC LTC 0.0017 18 5 100
#
# Running without input parameters will start Telegram bot listener
#
# Conservative - TP 1.13 / SL 0.95, volatile (NEO etc.) TP 1.18 / SL 0.94
# SL threshold applies both to original stop and to trailing stop

################################ Libraries ############################################
# Standard libraries 
import os
import time
import sys
from sys import exit, argv
from time import localtime, strftime
import subprocess   
import math
import smtplib
from email.MIMEMultipart import MIMEMultipart
from email.MIMEText import MIMEText
import urllib2
import decimal
from decimal import Decimal
from openpyxl import Workbook, load_workbook   
from openpyxl.styles import Font, Fill
import json # requests
from shutil import copyfile # to copy files
import numpy as np
import traceback

# Decimal precision and roubding 
decimal.getcontext().prec = 25
decimal.getcontext().rounding = 'ROUND_DOWN'

## Custom libraries 
from telegramlib import telegram # my lib to work with Telegram
from sqltools import query_lastrow_id, query # proper requests to sqlite db
from loglib import logfile # logging 
import platformlib as platform  # detecting the OS and assigning proper folders 

# Universal functions for all exchanges              
from exchange_func import getticker, getopenorders, cancel, getorderhistory, getorder, getbalance, selllimit, getorderbook, buylimit, getbalances, binance_price_precise, binance_quantity_precise, getpositions, closepositions, bitmex_leverage

# Using coinigy to get prices so that there are no stringent restrictions on api request rates (frequency)
from coinigylib import coinigy 
coinigy = coinigy()

################################ Config - part I ############################################

### Import a configuration file 
import config 

### TD analysis library
import tdlib as tdlib
td_info = tdlib.tdlib()

### Platform
platform = platform.platformlib()
platform_run, cmd_init, cmd_init_buy = platform.initialise() 
print "Initialising..."

### Set up the speedrun multiplier if need to test with higher speeds. 1 is normal, 2 is 2x faster 
speedrun = config.speedrun

### Telegram integration  
chat = telegram()

comm_method = config.comm_method 
send_messages = True

### Command prompt parameters  

### Default values
no_input = False 
trailing_stop_flag = True  # default mode is to have trailing stop

### Input parameters 
try: 
    simulation_param = argv[1]
    if simulation_param == 's': 
        simulation = True
        stop_loss = True
    elif simulation_param == 'r': 
        simulation = False
        stop_loss = True
    elif simulation_param == 'sns': 
        simulation = True
        stop_loss = False
    elif simulation_param == 'rns': 
        simulation = False
        stop_loss = False
    elif simulation_param == 'rnts':
        trailing_stop_flag = False 
        simulation = False
        stop_loss = True
    else: 
        no_input = True 

    exchange_abbr = argv[2].lower()
    if exchange_abbr not in ['btrx', 'bina', 'bmex']: 
        print 'Incorrect exchange specified (should be btrx, bina, or bmex)\n\n'
        exit(0)
    if exchange_abbr == 'btrx': 
        exchange = 'bittrex' 
        comission_rate = config.comission_rate_bittrex
    elif exchange_abbr == 'bina': 
        exchange = 'binance' 
        comission_rate = config.comission_rate_binance
    elif exchange_abbr == 'bmex': 
        exchange = 'bitmex' 
        comission_rate = config.comission_rate_bitmex
        
    market = argv[3].upper()
    try:
        trade, currency = market.split('-')
    except: 
        trade = market  # e.g. if only one market vs BTC is provided - such as XRPH18 on bitmex  
        currency = 'BTC'

    price_curr = float(argv[4])

    price_target = float(argv[5])
    sl_target = float(argv[6])
    price_entry = price_curr

    tp = round(price_target/price_curr, 5)
    sl = round(sl_target/price_curr, 5) 
    tp_p = (tp - 1.0)*100.0 
    sl_p = (1.0 - sl)*100.0 

    try:
        limit_sell_amount = float(argv[7])
    except: 
        limit_sell_amount = 0
    try:
        sell_portion = float(argv[8])
    except: 
        sell_portion = None    
    # print 'Trade', trade, 'currency', currency, 'simulation', simulation, 'price_curr', price_curr, 'tp', tp, 'sl', sl, limit_sell_amount, sell_portion  #DEBUG   
except:
    no_input = True 

# Terminate if there is no proper input    
if no_input:
    print '----------------------------------------------------------------------------------------------\n' + \
    'Run parameters not specified. Restart the script using:\n' + \
    'robot.py simulation (s/r/sns/rns) exchange basic_curr-altcoin entry_price TP SL [limit_of_amount_to_sell] [sell_portion]\n' +\
    'Example: > python robot.py s btrx BTC-LTC 0.0017 0.0021 0.0015 100\n\n' +\
    'Modes:\n>s (simulation with stop-loss)\n>r (real mode with stop-loss)\n>sns (simulation and stop only on profit)\n>rns (real and stop only on profit)'  
    exit(0) 
    
###  If simulation and parameters are not specified 
if simulation is True:
    if limit_sell_amount == 0: 
        limit_sell_amount = 100
    simulation_balance = limit_sell_amount
    sell_portion = limit_sell_amount

        
### Prices 
# something should be bought at a price_curr level to start from  

price_target = price_curr*tp
sl_target = price_curr*sl
price_entry = price_curr

#### Gmail login and pass (if used) 
fromaddr = config.fromaddr   # replace to a proper address 
toaddr = config.toaddr    # replace to a proper address 
email_passw = config.email_passw

################################ Config - part II ############################################
### Intervals and timers in seconds  

sleep_timer = config.sleep_timer                                    # Generic sleep timer. Applicable for the main monitoring loop and for the mooning procedure.
sleep_timer_buyback = config.sleep_timer_buyback      # Sleep timer for buybacks 
sleep_sale = config.sleep_sale                                      # Sleep timer for sell orders to be filled 
flash_crash_ind = config.flash_crash_ind                      # If something falls so much too fast - it is unusual and we should not sell (checking for 50% crashes)

## Interval and number of checks to get current (last) prices 
steps_ticker = config.steps_ticker  
sleep_ticker = config.sleep_ticker            

## Steps and timer for buybacks 
candle_steps = config.candle_steps         
candle_sleep = config.candle_sleep       

sleep_timer = int(sleep_timer/speedrun)
sleep_sale = int(sleep_sale/speedrun)
sleep_ticker = int(sleep_ticker/speedrun)
candle_steps = int(candle_steps/speedrun)

### To cancel buyback if there is an error and there were no sales made 
cancel_buyback = False 

### Bitmex margin 
bitmex_margin = config.bitmex_margin    # size of margin on bitmex, minor for now 

# Time analysis candles length 
td_period = config.td_period   # possible options are in line with ohlc (e.g. 1h, 4h, 1d, 3d); customisable. This sets up smaller time interval for dynamic stop losses and buy backs     
td_period_extended = config.td_period_extended   # possible options are in line with ohlc (e.g. 1h, 4h, 1d, 3d); customisable. This sets up larger time interval for buy backs (should be in line with the smaller one)         
td_period_ext_opposite = config.td_period_ext_opposite    # for buybacks in the different direction (e.g. initiating short after going long first) 

# Market reference for BTC 
btc_market_reference = config.btc_market_reference

### Starting variables  
main_curr_from_sell = 0     
commission_total = 0        
alt_sold_total = 0  
decrease_attempts_total = 0  
value_original = 0
contracts_start = 0 
stopped_mode = '' 
short_flag = False # whether we are shorting, applicable for bitmex 
bitmex_sell_avg = 0 # for bitmex price averaging 
price_flip = None # for the confirmation of stops on the previous candle (should be a price flip there to stop, on td_period). None by default 
price_exit = None 
sl_extreme = None 
sale_trigger = False  
market_ref = None      # if we refer to a different exchange and market for td stats 
exchange_abbr_ref = None 

### Handle the reference to a different set of prices (from finex) in the case of usd-btc and bitmex 
if market == 'USD-BTC' and exchange == 'bitmex' and btc_market_reference:        # put in the config 
    market_ref = config.market_ref
    exchange_abbr_ref = config.exchange_abbr_ref
    print "Reference market {} on {}".format(market_ref, exchange_abbr_ref) 

### Strategy and thresholds (for non-4H based action), contingency for TD (4h-based) action 
if currency in ['XMR', 'DASH', 'ETH', 'LTC', 'XMR']: 
    strategy = 'alt-med'
    diff_threshold = 0.045
elif currency == 'BTC': 
    strategy = 'btc'
    diff_threshold = 0.0255
else: 
    strategy = 'alt-volatile' 
    diff_threshold = 0.055

if strategy == 'btc': 
    var_contingency = 0.0086    # 0.86% lower/higher than the previous 4H candle for btc  # previously was 0.75%
else: 
    var_contingency = 0.01     # 1% lower/higher than the previous 4H candle for alts    

# Logger
logger = logfile(market, 'trade')

################################ Functions ############################################

### Log and print 
def lprint(arr):
    msg = ' '.join(map(lambda x: ''+ str(x), arr))
    try: 
        logger.write(msg)
        print msg
    except: 
        print 'Failed to print output due to the IO error'

    
##############################################  
##            Core get price / moon / sell functions           ##
##############################################

##################### Price comparison  
def strictly_increasing(L):
    return all(x<y for x, y in zip(L, L[1:]))
   
def equal_or_increasing(L):
    return all(x<=y for x, y in zip(L, L[1:]))

def strictly_decreasing(L):
    return all(x>y for x, y in zip(L, L[1:]))
    
def equal_or_decreasing(L):
    # Actually >= for our purposes
    return all(x>=y for x, y in zip(L, L[1:]))

##################### Processing sell outcome results and generating messages 
def process_stat(status): 
    global market
    global db, cur, job_id
    global cancel_buyback 
    
    flag = True   # default flag returned
    
    if status == 'stop':
        message = 'Finishing up normally'
        flag = False
        sql_string = "UPDATE jobs SET selling = 0 WHERE job_id = {}".format(job_id)     # DB update
        rows = query(sql_string)

    if status == 'err_low': 
        message = 'Trade amount was too small and returned error, finishing up'
        send_notification('Error: Too small trade', 'Too small trade to perform, finishing up')
        cancel_orders(market)
        flag = False
        cancel_buyback = True 
        
    if status == 'no_idea': 
        message = 'Sell calls did not return proper answer, aborting'
        send_notification('Error: No response from sell calls', 'Sell calls did not return proper answer, aborting')
        cancel_orders(market)
        flag = False
        cancel_buyback = True 
        
    if status == 'abort_telegram': 
        message = 'Aborted as requested via Telegram'
        cancel_orders(market)
        flag = False
        cancel_buyback = True 
        
    return flag, message

##################### Getting several last prices in short intervals instead of just one 
def get_last_price(market): 
    global exchange, exchange_abbr, coinigy
    
    ticker_upd = {}
    price_upd = 0
    failed_attempts = 0
    for i in range(1, steps_ticker + 1):
        try:
            #ticker_upd = getticker(exchange, market) 
            ticker_upd = coinigy.price(exchange_abbr, market)
            price_upd += ticker_upd
        except:
            # print "Issues with URL (!) for market", market
            failed_attempts += 1
            #time.sleep(sleep_ticker)
        
    # Logging failed attempts number
    if failed_attempts > 0: 
        lprint(["Failed attempts to receive price:", failed_attempts])    
        
    # If retreiving prices fails completely
    if failed_attempts == steps_ticker:     
        ticker_upd = None  
        try:
            send_notification('Maintenance', market + ' seems to be on an automatic maintenance. Will try every 5 minutes.')
        except: 
            lprint(["Failed to send notification"])    
        while ticker_upd is None: 
            time.sleep(300) # sleeping for 5 minutes and checking again
            lprint(["Market could be on maintenance. Sleeping for 5 minutes."])    
            try:
                #ticker_upd = getticker(exchange, market) 
                ticker_upd = coinigy.price(exchange_abbr, market)
            except: 
                ticker_upd = None
            price_upd = ticker_upd
    # If it is fine - get the average price 
    else: 
        price_upd = float(price_upd)/float(steps_ticker - failed_attempts)
        
    return price_upd

##################### Extreme in time series; returns value with the lowest or the highest ticker price among N-min intervals (candles) 
# type should be 'H' or 'L' (highest ore lowest in the series) 
def candle_extreme(type): 
    global exchange, exchange_abbr, market, candle_steps, candle_sleep
    global coinigy 
    ticker_upd = {}
    price_upd = 0
    price_extreme = 0
    failed_attempts = 0
    
    for i in range(1, candle_steps + 1): # 5 min: 100 checks x 3 sec (better indication than 30 checks x 10 sec); 80 x 3 for 4 minutes 
        try:
            #ticker_upd = getticker(exchange, market) 
            ticker_upd = coinigy.price(exchange_abbr, market)
            price_upd = ticker_upd
            if type == 'L': 
                if (price_extreme == 0) or (price_upd < price_extreme): 
                    price_extreme = price_upd
            if type == 'H': 
                if (price_extreme == 0) or (price_upd > price_extreme): 
                    price_extreme = price_upd
        except:
            #print "Issues with URL (!) for market", market
            failed_attempts += 1
        time.sleep(candle_sleep) 
        
    # Logging failed attempts number
    if failed_attempts > 0: 
        lprint(["Failed attempts to receive price:", failed_attempts])    
        
    # If retreiving prices fails completely
    if failed_attempts == steps_ticker:     
        ticker_upd = None # Change
        # Could be related to maintenance
        try:
            send_notification('Maintenance', market + ' seems to be on an automatic maintenance. Will try every 5 minutes.')
        except: 
            lprint(["Failed to send notification"])    
        while ticker_upd is None: 
            time.sleep(300)  
            lprint(["Market could be on maintenance. Sleeping for 5 minutes."])    
            try:
                #ticker_upd = getticker(exchange, market) 
                ticker_upd = coinigy.price(exchange_abbr, market)
            except: 
                ticker_upd = None
            price_upd = ticker_upd
            price_extreme = price_upd
            
    return price_extreme

##################### Candle analysis; returns high, low, and whether the price crossed a value - among N-min intervals (candles) 
def candle_analysis(cross_target): 
    global market, candle_steps, candle_sleep
    global exchange, exchange_abbr, coinigy 
    
    ticker_upd = {}
    price_upd = 0
    price_h = 0
    price_l = 0 
    crossed_flag = False
    failed_attempts = 0
    
    for i in range(1, candle_steps + 1): # 5 min: 100 checks x 3 sec (better indication than 30 checks x 10 sec) 
        try:
            #ticker_upd = getticker(exchange, market) 
            ticker_upd = coinigy.price(exchange_abbr, market) 
            price_upd = ticker_upd
            if (price_l == 0) or (price_upd < price_l): 
                price_l = price_upd
            if (price_h == 0) or (price_upd > price_h): 
                price_h = price_upd
            if price_upd >= cross_target: 
                crossed_flag = True 
        except:
            #print "Issues with URL (!) for market", market
            failed_attempts += 1
        time.sleep(candle_sleep) 
        
    # Logging failed attempts number
    if failed_attempts > 0: 
        lprint(["Failed attempts to receive price:", failed_attempts])    
    
    # If retreiving prices fails completely
    if failed_attempts == steps_ticker:     
        ticker_upd = None  
        # Could be related to maintenance
        try:
            send_notification('Maintenance', market + ' seems to be on an automatic maintenance. Will try every 5 minutes.')
        except: 
            lprint(["Failed to send notification"])    
        while ticker_upd is None: 
            time.sleep(300)  
            lprint(["Market could be on maintenance. Sleeping for 5 minutes."])    
            try:
                #ticker_upd = getticker(exchange, market) 
                ticker_upd = coinigy.price(exchange_abbr, market) 
            except: 
                ticker_upd = None
            price_upd = ticker_upd
            # If unsuccessful
            price_l = 0
            price_h = 0
    return price_l, price_h, crossed_flag
    
##################### Checking if we need to stop buyback
def check_bb_flag():
    global market 
    global db, cur, bb_id
    
    sell_initiate = False 
    sql_string = "SELECT abort_flag FROM bback WHERE id = {}".format(bb_id)
    rows = query(sql_string)
    try: 
        bb_flag = rows[0][0] # first result 
    except: 
        bb_flag = 0  
    return bool(bb_flag)

##################### Looking for rebuy points (buyback), based on 4H candles price action or simpler price action depending on data availability
def buy_back(price_base): 
    global bb_id, market, exchange_abbr, exchange, sleep_timer_buyback
    global td_data_available, start_time, bars, strategy, time_bb_initiated # bars actually need to be recalculated as 1h is used for buyback
    global short_flag, td_period, td_period_extended, td_period_ext_opposite
    global market_ref, exchange_abbr_ref
    
    direction = None # to return information on the direction of the new detected trend 
    bars_check_avail = None 

    ### Greetings (for logs readability) 
    lprint(["###################### BUY_BACK ###########################"])
    
    if strategy == 'btc': 
        diff_threshold = 0.005 # threshold as a low/high +- 0.5% depending on the direction 
    else: 
        diff_threshold = 0.01 # 1% for alts 
    
    flag_reb_c = True 
    td_first_run = True 
    bback_result = False 

    if td_data_available != True: ## using a simple 5-min candles analysis if there is no 4H price data 
        # print "Base: ", price_base    #DEBUG 
        price_l_arr = np.zeros(5)        #5x5-min candlesticks
        price_h_arr = np.zeros(5)       #5x5-min candlesticks
        crossed_arr = np.bool(5)        # for crossed
        lprint([market, ': filling the price array'])
        
        bback_result_long = False 
        bback_result_short = False 
        
        # Filling the prices array with 5x5min candles
        while 0 in price_l_arr: 
            price_l, price_h, crossed_flag = candle_analysis(price_base)
            price_l_arr = np.append(price_l_arr, price_l)
            price_h_arr = np.append(price_l_arr, price_l)
            crossed_arr = np.append(crossed_arr, crossed_flag)
            price_l_arr = np.delete(price_l_arr, [0])
            price_h_arr = np.delete(price_h_arr, [0])
            crossed_arr = np.delete(crossed_arr, [0])
            # print "Lows", price_l_arr, '\nHighs', price_h_arr, '\nCrosses', crossed_arr   #DEBUG

        # Running until need to cancel 
        while flag_reb_c: 
            crossed_conf = (True in crossed_arr)                            # Any of candles should cross price_base   
            # LONGS check 
            lows_conf = equal_or_increasing(price_l_arr)                 # Higher or equal lows
            num_conf_long = ((price_h_arr >= price_base).sum()) >= 3     # At least 3 of 5 candles highs should be equal or above x
            bback_result_long = bool(lows_conf * crossed_conf * num_conf_long)
            lprint([market, ": base", price_base, '| lows holding or higher', lows_conf, '| highs lower than base confirmation:', num_conf_long, '| crossed flag:', crossed_conf, '| result (long):', bback_result_long])
            # SHORTS check
            highs_conf = equal_or_decreasing(price_l_arr)                 # Lower or equal highs
            num_conf_short = ((price_l_arr <= price_base).sum()) >= 3     # At least 3 of 5 candles lows should be equal or below x
            bback_result_short = bool(highs_conf * crossed_conf * num_conf_short)
            lprint([market, ": base", price_base, '| highs holding or lower', highs_conf, '| highs lower than base confirmation:', num_conf_short, '| crossed flag:', crossed_conf, '| result (short):', bback_result_short])
            
            # Check if we need to cancel 
            stop_bback = check_bb_flag()
            if stop_bback: 
                bback_result = False 
                flag_reb_c = False 
            
            # If we need to exit to proceed with buyback
            if bback_result_long: 
                bback_result = True
                direction = 'up' 
                lprint([market, ": initiating buyback"])
                flag_reb_c = False 
            if bback_result_short: 
                bback_result = True
                direction = 'down' 
                lprint([market, ": initiating buyback"])
                flag_reb_c = False 
                
            # Get new values 
            price_l, price_h, crossed_flag = candle_analysis(price_base)
            price_l_arr = np.append(price_l_arr, price_l)
            price_h_arr = np.append(price_l_arr, price_l)
            crossed_arr = np.append(crossed_arr, crossed_flag)
            price_l_arr = np.delete(price_l_arr, [0])
            price_h_arr = np.delete(price_h_arr, [0])
            crossed_arr = np.delete(crossed_arr, [0])
            
            # Updating DB
            if bb_id is not None: 
                sql_string = "UPDATE bback SET curr_price = {} WHERE id = {}".format(price_h, bb_id) 
                rows = query(sql_string)
               
            # Sleeping 
            time.sleep(sleep_timer_buyback)     
            
    ## If there is detailed 4H (or larger interval) data available 
    else: 
        # Update to set stops according to 4H candles and TD 
        if td_first_run: 
            time_hour = time.strftime("%H")
            
        while flag_reb_c: 
            # Checking the need to update 
            time_hour_update = time.strftime("%H")
            if (time_hour_update <> time_hour) or td_first_run:
                # If this is the first run 
                if td_first_run: 
                    td_first_run = False 

                # Updating time 
                time_hour = time_hour_update
                # Updating TD values 
                bars = td_info.stats(market, exchange_abbr, td_period, 50000, 5, short_flag, market_ref, exchange_abbr_ref)     
                try: 
                    bars_extended = td_info.stats(market, exchange_abbr, td_period_extended, 80000, 5, short_flag, market_ref, exchange_abbr_ref)   
                    bars_check_avail = True 
                except: 
                    bars_check_avail = False 
                try: 
                    bars_ext_opposite = td_info.stats(market, exchange_abbr, td_period_ext_opposite, 80000, 10, short_flag, market_ref, exchange_abbr_ref)   
                    bars_check_avail = True 
                except: 
                    bars_check_avail = False 
                    
            # Check if we need to cancel 
            stop_bback = check_bb_flag()
            if stop_bback: 
                bback_result = False 
                flag_reb_c = False 
            
            # Checking time elapsed from the start of buyback 
            time_elapsed = (math.ceil(time.time() - time_bb_initiated ))/60    
            
            # Getting the current price and showing info on potential longs or potential shorts 
            price_upd = get_last_price(market)
            if bars['td_direction'].iloc[-2] == 'up': #LONGS potential 
                lprint([  exchange, market, "TD setup:", bars['td_setup'].iloc[-2], "| TD direction:", bars['td_direction'].iloc[-2], "4H candle high:", bars['high'].iloc[-2], "Current price:", price_upd, "Time elapsed (min):", time_elapsed  ])    
                if bars_check_avail: 
                    lprint([  exchange, market, "TD setup (extended):", bars_extended['td_setup'].iloc[-2], "| TD direction:", bars_extended['td_direction'].iloc[-2] ])    
                else: 
                    lprint(["Extended timeframe price data unavailable"])    
            elif (bars['td_direction'].iloc[-2] == 'down') and (exchange == 'bitmex'): #SHORTS potential, only for bitmex 
                lprint([  exchange, market, "TD setup:", bars['td_setup'].iloc[-2], "| TD direction:", bars['td_direction'].iloc[-2], "4H candle low:", bars['low'].iloc[-2], "Current price:", price_upd, "Time elapsed (min):", time_elapsed  ])    
                if bars_check_avail: 
                    lprint([  exchange, market, "TD setup (extended):", bars_extended['td_setup'].iloc[-2], "| TD direction:", bars_extended['td_direction'].iloc[-2] ])    
                else: 
                    lprint(["Extended timeframe price data unavailable"])    
                    
            # Updating DB
            if bb_id is not None: 
                sql_string = "UPDATE bback SET curr_price = {} WHERE id = {}".format(price_upd, bb_id) 
                rows = query(sql_string)
            
            # Checking if we should buy back 
            # Check longs 
            if (bars['td_direction'].iloc[-2] == 'up') and (bars['td_direction'].iloc[-1] == 'up') and (time_elapsed > 60) and (price_upd > (bars['high'].iloc[-2])*(1 + diff_threshold)):      # switching to long    
                # Depending on the short flag 
                if not short_flag:  # same direction - checking bars_extended 
                    bars_check = bars_extended
                else:   # different direction - checking a larger period 
                    bars_check = bars_ext_opposite
                # Checking the conditions                 
                if (bars_check_avail and bars_check['td_direction'].iloc[-1] == 'up') or (bars_check_avail == False): 
                    bback_result = True 
                    direction = 'up'
                    flag_reb_c = False 
                    lprint(["TD buyback initiated on the long side"])
                    if bars_check_avail == False: 
                        lprint(["Note that higher - timeframe TD analysis is not available"])    
            
            # Check shorts  
            if (bars['td_direction'].iloc[-2] == 'down') and (bars['td_direction'].iloc[-1] == 'down') and (time_elapsed > 60) and (price_upd < (bars['low'].iloc[-2])*(1 - diff_threshold)) and (exchange == 'bitmex'):     # switching to short, only for bitmex       
                # Depending on the short flag 
                if short_flag:  # same direction - checking bars_extended 
                    bars_check = bars_extended
                else:     # different direction - checking a larger period 
                    bars_check = bars_ext_opposite
                # Checking the conditions      
                if (bars_check_avail and bars_check_avail['td_direction'].iloc[-1] == 'down') or (bars_check_avail == False): 
                    bback_result = True 
                    direction = 'down'
                    flag_reb_c = False 
                    lprint(["TD buyback initiated on the short side"])
                    if bars_check_avail == False: 
                        lprint(["Note that higher - timeframe TD analysis is not available"])    
            
            # Sleeping 
            time.sleep(sleep_timer_buyback)     
            
    # Finishing up 
    return bback_result, direction
    
##################### Cancelling active orders on particular market if there are any 
def cancel_orders(market):    
    my_orders = getopenorders(exchange, market)
    # print "Orders", my_orders #DEBUG 
    if my_orders <> '': 
        for val in my_orders:
            lprint(["Cancelling open order:", val['OrderUuid'], "quantity", val['Quantity'], 
                   "quantity remaining", val['QuantityRemaining'], "limit", val['Limit'], "price", val['Price']
                   ])
            cancel_stat = cancel(exchange, market, val['OrderUuid'])
            
            # Wait for a moment if needed 
            # time.sleep(1)

##################### Update information on performed orders
def sell_orders_info():
    global simulation, main_curr_from_sell, commission_total, alt_sold_total, orders_start, no_sell_orders, market, limit_sell_amount
    global exchange 
    global price_exit, bitmex_sell_avg
    
    # Updating order history to collect information on new orders (disable in the simulation mode)
    # Further speed improvement would be to change the structure to a proper dict here right away    
    
    try: # to handle errors 
        if simulation != True: 
            # Reset values if we are not simulating
            main_curr_from_sell = 0     
            commission_total = 0        
            alt_sold_total = 0 
            
            # Getting information on _sell_ orders executed
            orders_opening_upd = getorderhistory(exchange, market) 
            for elem in orders_opening_upd: 
                orders_new.add(elem['OrderUuid'])
            orders_executed = orders_new.symmetric_difference(orders_start) 
     
            if orders_executed == set([]):
                lprint(["No sell orders executed"])
                no_sell_orders = True 
            else:
                lprint(["New executed orders"])  
                
                for elem in orders_executed: 
                    order_info = getorder(exchange, market, elem)               
                    ''' # Does not work due to data limitations - revise 
                    if exchange == 'bitmex':         
                        if order_info['Status'] != 'Canceled': 
                            main_curr_from_sell += float(order_info['simpleCumQty'])    # this does not work, revise (all orders are reflected with 0 quantity remaining) 
                        commission_total += 0 
                        qty_sold = order_info['Quantity'] - order_info['QuantityRemaining'] 
                        alt_sold_total += qty_sold             
                        lprint([">", elem, "price", order_info['Price'], "quantity sold", qty_sold, "BTC from sale", float(order_info['simpleCumQty']) ]) #DEBUG 
                    else: 
                        main_curr_from_sell += order_info['Price']  
                        commission_total += order_info['CommissionPaid']
                        qty_sold = order_info['Quantity'] - order_info['QuantityRemaining'] 
                        alt_sold_total += qty_sold             
                        lprint([">", elem, "price", order_info['Price'], "quantity sold", qty_sold ]) #DEBUG 
                    ''' 
                    if exchange != 'bitmex': 
                        main_curr_from_sell += order_info['Price']  
                        commission_total += order_info['CommissionPaid']
                        qty_sold = order_info['Quantity'] - order_info['QuantityRemaining'] 
                        alt_sold_total += qty_sold                            
                        lprint([">", elem, "price", order_info['Price'], "quantity sold", qty_sold ]) #DEBUG 
                    else: 
                        price_exit = bitmex_sell_avg
                        if price_exit != 0: 
                            if market == 'USD-BTC': 
                                main_curr_from_sell = contracts_start/price_exit
                            else: 
                                main_curr_from_sell = contracts_start*price_exit
                lprint(["Total price", main_curr_from_sell, "alts sold total", alt_sold_total]) #DEBUG
        else:
            # If the simulation is True - main_curr_from_sell will have simulated value and the commission would be zero. Updating quantity. 
            alt_sold_total = limit_sell_amount
            price_exit = get_last_price(market)
            if exchange == 'bitmex': 
                if market == 'USD-BTC': 
                    main_curr_from_sell = contracts_start/price_exit
                else: 
                    main_curr_from_sell = contracts_start*price_exit

    except: 
        err_msg = traceback.format_exc()
        comm_string = 'Could not het sell orders history from {} on {}. Reason: {}. Check the results'.format(market, exchange, err_msg)
        lprint([comm_string])    
        chat.send(comm_string)  
    
##################### Sell orders outcome 
def sell_orders_outcome():
    global no_sell_orders, total_gained, main_curr_from_sell, value_original, commission_total, total_gained_perc, market
    global price_exit, contracts_start # to use in buyback
    
    if no_sell_orders != True: 
        # Calculating totals 
        total_gained = float(main_curr_from_sell) - float(value_original) - float(commission_total)
        
        # Here division by zero error handling
        if float(value_original)  != 0: 
            total_gained_perc = 100*float(total_gained)/float(value_original)   
        else: 
            total_gained_perc = 0 
        
        # Depending on the trade direction 
        if (short_flag and (total_gained_perc < 0)) or (not short_flag and (total_gained_perc >= 0)): 
            txt_result = 'gained'
        else: 
            txt_result = 'lost'  
        
        # Average exit price (value/quantity)
        if exchange != 'bitmex':   
            price_exit = float(main_curr_from_sell)/float(alt_sold_total)
        
        ''' 
        # Bitmex: Does not work due to data limitations - revise 
        # price_exit = float(contracts_start)/float(main_curr_from_sell)   # for bitmex, calculation is done through contracts    # commented until further fix
        ''' 
        print  "Price exit calc: price_exit {}, contracts_start {}, main_curr_from_sell {}".format(price_exit, contracts_start, main_curr_from_sell) # DEBUG 
            
        percent_gained = str(round(total_gained_perc, 2))
        trade_time = strftime("%Y-%m-%d %H:%M", localtime())
        
        lprint(['Total from all sales', main_curr_from_sell, 'total commission', commission_total])
        lprint(['Profit ', total_gained, ':', round(total_gained_perc, 2), '%']) 
        
        # Send the notification about results    
        # To improve: if bitmex is used, margin should be accounted for 
        msg_result = '{}: Total {} gained from all sales: {}. Commission paid: {}. Trade outcome: {} % {}. \nEntry price: {}, exit price: {}'.format(market, str(trade), main_curr_from_sell, str(commission_total),  percent_gained , txt_result, str(price_entry), str(price_exit))
        send_notification('Finished', msg_result) 
        
        # Update the xls register 
        try:
            wb = load_workbook(config.trade_hist_filename)
            ws = wb['BOT']
            new_line = [trade_time, trade, currency, alt_sold_total, price_curr, price_exit, main_curr_from_sell, total_gained, percent_gained, simulation]
            ws.append(new_line)
            max_row = ws.max_row
            # Apply a style 
            index_row = "{}:{}".format(max_row, max_row) 
            for cell in ws[index_row]:
                cell.font = Font(name='Arial', size=10)
            wb.save(config.trade_hist_filename)
            
            #if platform_run != 'Windows':  #uncomment if needed 
            #    copyfile('/home/illi4/Robot/Trade_history.xlsx', '/mnt/hgfs/Shared_folder/Trade_history.xlsx')
            
        except: 
            lprint(['Trade history xls unavailable']) 

##################### Setting stop loss based on price data
def stop_reconfigure(mode = None): 
    global db, cur, job_id
    global time_hour
    global market, exchange_abbr, strategy 
    global price_entry, short_flag, td_period
    global var_contingency
    global bars_4h 
    global market_ref, exchange_abbr_ref
    
    price_flip_upd = None # default is None   
    price_direction_move = None 
    sl_target_upd = None 
    sl_upd = None 
    sl_p_upd = None  
    sl_extreme_upd = None # for the absolute min / max of TD setup 
    
    time_hour_update = time.strftime("%H")
    if (time_hour_update <> time_hour) or mode == 'now': 
        # Updating the current hour and the TD values 
        time_hour = time_hour_update
        bars_4h = td_info.stats(market, exchange_abbr, td_period, 50000, 5, short_flag, market_ref, exchange_abbr_ref)     

        ''' 
        # Checking whether there was a price flip - previous logic 
        if bars_4h['td_direction'].iloc[-1] != bars_4h['td_direction'].iloc[-2]: 
            price_flip_upd = True 
        else: 
            price_flip_upd = False 
        ''' 
        # New logic: return the TD direction of the last candle per td_interval 
        price_direction_move = bars_4h['td_direction'].iloc[-1]      # return 'up' or 'down' 
        price_direction_move_previous = bars_4h['td_direction'].iloc[-2]      # return 'up' or 'down' 
        #print "CHECK: short flag", short_flag, "price_direction", price_direction_move   #DEBUG
        
        # We will be considering that there is a price flip if we have a candle in setup with different colour which is followed by the same colour 
        if ((not short_flag and price_direction_move == 'down' and price_direction_move_previous == 'down') 
        or (short_flag and price_direction_move == 'up' and price_direction_move_previous == 'up')): 
            price_flip_upd = True 
            #print ">>>>>> Price_flip_upd", price_flip_upd #DEBUG 

        if not short_flag: # the position is long 
            sl_target_upd = bars_4h['low'].iloc[-2] * (1 - var_contingency)   
            sl_upd = round(sl_target_upd/price_entry , 5) 
            sl_p_upd = (1.0 - sl_upd)*100.0 
            if bars_4h['move_extreme'].iloc[-1]  is not None: 
                sl_extreme_upd = bars_4h['move_extreme'].iloc[-1] * (1 - var_contingency)      
        else: # the position is short 
            sl_target_upd = bars_4h['high'].iloc[-2] * (1 + var_contingency)   
            sl_upd = round(sl_target_upd/price_entry , 5) 
            sl_p_upd = (1.0 + sl_upd)*100.0  
            if bars_4h['move_extreme'].iloc[-1]  is not None: 
                sl_extreme_upd = bars_4h['move_extreme'].iloc[-1] * (1 + var_contingency)     
        
        lprint([  "New stop loss level based on the last candle: {}, setup direction: {}. Flip: {}".format(sl_target_upd, price_direction_move, price_flip_upd) ])
        lprint([  "New extreme stop value:", sl_extreme_upd ])
        #print ">>> Returning price_flip_upd {}, sl_target_upd {}, sl_upd {}, sl_p_upd {}, sl_extreme_upd {}".format(price_flip_upd, sl_target_upd, sl_upd, sl_p_upd, sl_extreme_upd)  #DEBUG
        
        ''' #Old code
        if short_flag != True: # LONGS  
            if bars_4h['td_direction'].iloc[-1] == 'up': 
                sl_target_upd = bars_4h['low'].iloc[-1] * (1 - var_contingency)   
                sl_upd = round(sl_target_upd/price_entry , 5) 
                sl_p_upd = (1.0 - sl_upd)*100.0 
                lprint(["New stop loss level based on bullish TD:", sl_target_upd, "| price direction:", price_direction])
            else:    
                lprint(["No bullish 4H candle to update the stop loss"])  
        else:  # SHORTS   
            if bars_4h['td_direction'].iloc[-1] == 'down': 
                sl_target_upd = bars_4h['high'].iloc[-1] * (1 + var_contingency)   
                sl_upd = round(sl_target_upd/price_entry , 5) 
                sl_p_upd = (1.0 + sl_upd)*100.0 
                lprint(["New stop loss level based on bearish TD:", sl_target_upd, "| price direction:", price_direction])
            else:    
                lprint(["No bearish 4H candle to update the stop loss"])  
        ''' 
        
    # Updating the db with the current SL value 
    if sl_target_upd is not None: 
        sql_string = "UPDATE jobs SET sl={}, sl_p={} WHERE job_id={}".format(sl_target_upd, sl_upd, job_id)
        rows = query(sql_string)   
    
    return price_flip_upd, sl_target_upd, sl_upd, sl_p_upd, sl_extreme_upd
            
            
##################### Mooning trajectory procedure
##################### Currently works in the same way as just the main cycle when TD price data is available (simply reconfiguring stops dynamically)        
def to_the_moon(price_reached):     
    # Global variables used 
    global main_curr_from_sell, value_original, price_curr, commission_total, price_target, t_m_id, approved_flag, offset_check, comission_rate
    global sleep_timer
    global db, cur, job_id
    global stopped_price
    global trailing_stop_flag, start_time, bars, strategy, diff_threshold
    global sl, sl_target, sl_p
    global short_flag, price_flip, sl_extreme
    global bars_4h

    sale_trigger = False # default
    
    # Thresholds for post-profit fallback for BTC or ALTS, when detailed price data is not available 
    if market == 'USDT-BTC': 
        post_sl_level = 0.9745     # Fix of -2.55% for BTC
    else: 
        post_sl_level = 0.975    # Fix of -2.5% for all the post-profit cases (alts)  
        
    price_max = price_reached       # this will be changed mooning forward
    price_cutoff = price_reached * post_sl_level   # to sell on original TP if we fall below 
    if td_data_available: 
        trailing_stop = sl_target 
    else: 
        trailing_stop = price_max * post_sl_level    # to sell on new high * stop loss threshold   
    
    lprint(["Mooning from:", price_max])   
    rocket_flag = True
    
    # Running the loop 
    while rocket_flag:  
        # Update to set stops according to 4H candles and TD 
        if td_data_available: 
            price_flip_upd, sl_target_upd, sl_upd, sl_p_upd, sl_extreme_upd = stop_reconfigure()
            #print ">>> Returned price_flip {}, sl_target_upd {}, sl_upd {}, sl_p_upd {}, sl_extreme_upd {}".format(price_flip_upd, sl_target_upd, sl_upd, sl_p_upd, sl_extreme_upd)  #DEBUG
            if sl_target_upd is not None: 
                trailing_stop = sl_target_upd
                sl = sl_upd
                sl_p = sl_p_upd    
            if sl_extreme_upd is not None: 
                sl_extreme = sl_extreme_upd
            if price_flip_upd is not None: 
                price_flip = price_flip_upd
            
        price_last_moon = get_last_price(market)
        increase_info = 100*float(price_last_moon - price_target)/float(price_target) 
        lprint(["Price update:", price_last_moon, "in comparison with the original target:", round(increase_info, 2), "%"])

        # Updating the db 
        sql_string = "UPDATE jobs SET price_curr={}, percent_of={}, mooning={} WHERE job_id={}".format(round(price_last_moon, 8), str(round(increase_info, 2)), 1, job_id)
        rows = query(sql_string)
        
        # Depending on whether there is short or long 
        if ((short_flag != True) and (price_last_moon > price_max)) or ((short_flag == True) and (price_last_moon < price_max)):  
            # Setting higher thresholds if there is no 4H data
            price_max = price_last_moon
            if not td_data_available: 
                trailing_stop = price_max * post_sl_level        
            lprint(["Last price:", price_max, "| trailing stop", trailing_stop, "| original take profit", price_cutoff])

        #  Checking if this is a time to sell now   
        #  starting only when trailing_stop_flag is active (should not be doing this for BTC runs) 
        # print ">> Price last moon (to compare)", price_last_moon, "maximum price", price_max, "price_cutoff", price_cutoff, "trailing_stop", trailing_stop  # DEBUG # 
        
        if trailing_stop_flag: 
            # Simplified this back to basics as we are using the 4H rule and selling if we are falling behind the bullish candle 
            # Depending on the long or short 
            if (((not short_flag) and price_flip and (price_last_moon <= min(price_cutoff, trailing_stop)) )  # if we are long and the price drops below original or trailing stop 
            or (short_flag and price_flip and (price_last_moon >= max(price_cutoff, trailing_stop))) ):  
                lprint(["Run out of fuel @", price_last_moon])
                # Check if we need to sell. No need to do this if we have price action data (backtested for performance) 
                if not td_data_available: 
                    sale_trigger = ensure_sale(price_last_moon)   
                else: 
                    sale_trigger = True 
                lprint(["Sale trigger (post-profit)", sale_trigger])
            
            # Also check for extreme moves out of the TD setup 
            if sl_extreme is not None: 
                if (short_flag and (price_last_moon > sl_extreme)) or (not short_flag and (price_last_moon < sl_extreme)): 
                    lprint(["Breached TD extreme stop", price_last_moon])
                    # Check if we need to sell
                    if not td_data_available: 
                        sale_trigger = ensure_sale(price_last_moon)   
                    else: 
                        sale_trigger = True 
                    lprint(["Sale trigger", sale_trigger])
            
            ''' 
            # Disabled to aim at longer time frames and time analysis 
            # It is a good idea to sell at +45% when something is pumping 
            if increase_info > 45:
                sale_trigger = True 
            ''' 
            
            # Now checking sale trigger and selling if required         
            if sale_trigger == True:  
                lprint(["Triggering trailing stop on", price_last])
                send_notification('Sell: Post-TP', exchange + ' : ' + market + ': Triggering trailing stop on the level of ' + str(price_last))
                status = sell_now(price_last_moon)
                # Update the status
                rocket_flag, stat_msg = process_stat(status)
                lprint([stat_msg])            
                # For buyback - using rebuy price
                if short_flag:  
                    stopped_price = min(price_cutoff, trailing_stop)
                else: 
                    stopped_price = max(price_cutoff, trailing_stop)
                            
        # Check if 'sell now' request has been initiated
        sell_init_flag = check_sell_flag()
        if sell_init_flag == True:       
            lprint(["Sale initiated via Telegram @", price_last])
            status = sell_now(price_last_moon)
            sql_string = "UPDATE jobs SET selling = 0 WHERE job_id = {}".format(job_id)     # updating the DB 
            rows = query(sql_string)
            
            # Handling results
            rocket_flag, stat_msg = process_stat(status)
            lprint([stat_msg])
            # For buyback - using rebuy price
            if short_flag:  
                stopped_price = min(price_cutoff, trailing_stop)
            else: 
                stopped_price = max(price_cutoff, trailing_stop)
                
        # Checking Telegram requests and answering 
        if rocket_flag:
            approved_flag = check_cancel_flag()
            if approved_flag == False: 
                lprint(["Shutdown was requested via Telegram"])   
                sleep_timer = 0
            time.sleep(sleep_timer)

        if approved_flag == False:  # aborting if asked          
            status = 'abort_telegram'
            rocket_flag, stat_msg = process_stat('abort_telegram')

    # Finished the loop - returning the proper code
    return status

##################### Anti-manipulation and anti-flash-crash filter for cases when we do not rely on time analysis (there is no data) 
def ensure_sale(check_price): 
    global short_flag
    
    proceed_sale = False           # default 
    price_arr = np.zeros(3)         # 3 * N-min candlesticks  (see candle_extreme for N) 
    lprint(["Running ensure_sale check"])
    
    ## Filling the prices array - will be checking for lower highs 
    while (0 in price_arr):  
        approved_flag = check_cancel_flag() # checking Telegram requests and answering 
        if approved_flag == False: 
            break
    
        price_lowest, price_highest, crossed_flag_info = candle_analysis(check_price)           #candle_extreme('H')  
        if short_flag != True: # LONGS  
            price_arr = np.append(price_arr, price_highest)
        else:  # SHORTS   
            price_arr = np.append(price_arr, price_lowest)
        price_arr = np.delete(price_arr, [0])
        
        # Selling on the series of lower or same highs of 3 x N-min candlesticks when the price array is filled for longs: 
        if short_flag != True: # LONGS  
            if (0 not in price_arr): 
                lprint(["High in the candle:", price_highest, "| lower or same highs:", equal_or_decreasing(price_arr)])  #lprint([price_arr]) # DEBUG
                if equal_or_decreasing(price_arr): 
                    proceed_sale = True
                    break
            else: 
                lprint(["High in the candle:", price_highest])  #lprint([price_arr]) # DEBUG
                
            # If we are back above the check_price value - exit the cycle and return false 
            if price_highest > check_price: 
                lprint(["Cancelling ensure_sale since the price is back to normal"])  
                proceed_sale = False
                break
                
        else: # SHORTS      
            if (0 not in price_arr): 
                lprint(["Low in the candle:", price_lowest, "| higher or same lows:", equal_or_increasing(price_arr)])  #lprint([price_arr]) # DEBUG
                if equal_or_increasing(price_arr): 
                    proceed_sale = True
                    break
            else: 
                lprint(["Low in the candle:", price_lowest])  #lprint([price_arr]) # DEBUG
                
            # If we are back above the check_price value - exit the cycle and return false 
            if price_lowest < check_price: 
                lprint(["Cancelling ensure_sale since the price is back to normal"])  
                proceed_sale = False
                break
                
 
    '''
    # Not useful really considering the approach above
    if short_flag != True: # LONGS  
        if (price_arr.min() < (price_curr * flash_crash_ind)) and (0 not in price_arr):
            lprint(["Ridiculously low price, better check it."])
            chat.send(market +": ridiculously low price, better check it")
            proceed_sale = False 
    else: # SHORTS   
        if (price_arr.max() > (price_curr * (1 - flash_crash_ind))) and (0 not in price_arr):
            lprint(["Ridiculously high price, better check it."])
            chat.send(market +": ridiculously high price, better check it")
            proceed_sale = False 
    ''' 
    return proceed_sale     

##################### Main sell function to sell at current prices   
# Will be performed until the balance available for sale is zero or slightly more      
def sell_now(at_price):

    global bitmex_sell_avg  # for bitmex average price calc 
    bitmex_sell_avg_arr = []
    
    # To decrease price gradually compared to the last average sell price if orders are not filled. Start with zero (percent), maximum 5%
    decrease_price_step = 0.0 
    decrease_attempts_total = 0 
    # First run flag now to sleep on the first call 
    proceed_w_sleep = False
    
    # Global variables used 
    global main_curr_from_sell, value_original, price_curr, commission_total, simulation, currency, market, t_m_id, approved_flag, offset_check, simulation_balance, sell_portion, limit_sell_amount, comission_rate, exchange
    global sleep_sale, steps_ticker, sleep_ticker
    global db, cur, job_id
    global chat
    global balance_start, contracts_start, short_flag
    
    # Starting balance for further use. Should be done with all orders cancelled
    cancel_orders(market)
    
    # Get balance
    if simulation != True: 
        balance = getbalance(exchange, currency)
        balance_start  = Decimal('{0:.8f}'.format(balance['Available']))   # to correctly work with decimal numbers; not needed for bitmex 

        if exchange != 'bitmex':         
            lprint(["Balance available to sell", balance_start])    #DEBUG
    
    if limit_sell_amount is not None: 
        limit_sell_amount = Decimal(str(limit_sell_amount))     # using str, we will not have more decimal numbers than needed
    if sell_portion is not None: 
        sell_portion = Decimal(str(sell_portion))  
    
    if simulation == True: 
        balance_start = Decimal(str(simulation_balance))
        balance_available = Decimal(str(simulation_balance))
        remaining_sell_balance = Decimal(str(simulation_balance))
        
    # Limiting if required. Should be done with orders cancelled
    if (limit_sell_amount < balance_start) and (limit_sell_amount > 0):
        balance_adjust = Decimal(str(balance_start)) - Decimal(str(limit_sell_amount))
        balance_start = Decimal(str(limit_sell_amount))
        #print ">> Adjust", balance_adjust, "Bal_start", balance_start, "Limit sell am", limit_sell_amount      #DEBUG 
        lprint(["Limiting total amount to be sold. Total:", limit_sell_amount, "Adjustment:", balance_adjust])
    else:
        balance_adjust = 0

    # For bitmex, we will be trading contracts, no adjustments are available. Getting the balances and setting the original value 
    if exchange == 'bitmex': 
        if simulation != True: 
            # There were issues with testnet returning blanks so changed this 
            contracts_check = {}
            positions = getpositions(exchange, market)  # first not empty result 
            for position in positions: 
                if position != {}: 
                    contracts_check = position 
                    break # exit the for loop 
            print 'contracts_check', contracts_check #TEST 
            # If nothing was found  
            if contracts_check == {}: 
                sell_run_flag = False
                contracts = 0
            else: 
                if market == 'USD-BTC': 
                    contracts = contracts_check['contracts'] 
                    value_original = Decimal(str(contracts_check['contracts_no']))
                else: 
                    contracts = contracts_check['contracts_no'] 
                    value_original = Decimal(str(contracts))*Decimal(price_entry) #HERE
               
                contracts_start = contracts
                balance_available = contracts
                balance_adjust = 0 
                sell_portion = balance_available
             
        else: # if we are in the simulation mode 
            contracts =  price_entry * simulation_balance    #get_last_price(market) * simulation_balance
            contracts_start = contracts
            value_original = simulation_balance
    else: # for other exchanges     
        value_original = Decimal(str(price_entry)) * balance_start    
 
    lprint(["Original value:", value_original])
    
    # Main sell loop
    sell_run_flag = True
    stopmessage = 'stop' # default stop message meaning successful sale
    
    while sell_run_flag: 
        decrease_price_flag = False     # Reset the overall flag to decrease price 
      
        # Wait until existing orders are cancelled - that is why we need sleep here and not in the end 
        # Checking Telegram requests and cancelling if needed
        if proceed_w_sleep: 
            time.sleep(sleep_sale)
        
        # 0. Check open orders, cancel if unfilled, and decrease price further compared to average last 
        my_orders = getopenorders(exchange, market)
        if my_orders <> '': 
            for val in my_orders:
                # Checking if some are open not filling
                if (val['Quantity'] == 0):
                    unfilled_prop = 0
                else:
                    unfilled_prop = Decimal(str(val['QuantityRemaining']))/Decimal(str(val['Quantity']))
                if unfilled_prop >= 0.05:  # if more than 5% still left in the order
                    lprint(["Cancelling unfilled order:", val['OrderUuid'], "quantity", val['Quantity'], 
                           "quantity remaining", val['QuantityRemaining'], "limit", val['Limit'], "price", val['Price']
                           ]) 
                    cancel_stat = cancel(exchange, market, val['OrderUuid'])
                    time.sleep(5) # Wait for cancellations to be processed just in case 
                    # Then we will get information on available balance which includes cancellations
                    # Set decrease price flag
                    decrease_price_flag = True
                    decrease_attempts_total += 1

        # Decrease price more compared to last prices if required
        if (decrease_price_step < 0.05) and decrease_price_flag:
            if short_flag != True: #LONG
                decrease_price_step += 0.005
                lprint(["Sell price will be decreased on", decrease_price_step*100, "%"]) 
            else: #SHORT 
                decrease_price_step -= 0.005
                lprint(["Sell price will be increased on", decrease_price_step*100, "%"]) 
            
        # Notify if a position cannot be sold for a long time 
        if decrease_attempts_total >= 30: 
            time_passed = int(decrease_attempts_total*(sleep_sale + steps_ticker*sleep_ticker)/60)
            lprint(["Unable to sell the position for more than", time_passed, "minutes"]) 
            chat.send(market +": unable to sell the position for more than " + time_passed + " minutes")
                        
        # 1. Get the available balance and proceed with selling       
        if simulation != True: 
            balance = getbalance(exchange, currency)
            balance_available = Decimal('{0:.8f}'.format(balance['Available']))
            # print ">> Balance_available", balance_available #DEBUG 
        else:
            # If we are in the simulation mode - use the value from the previous run
            balance_available = remaining_sell_balance           
        
        # For bitmex, we will be trading contracts, no adjustments are available 
        if exchange == 'bitmex': 
            # There were issues with testnet returning blanks so changed this 
            contracts_check = {}
            positions = getpositions(exchange, market)  # first not empty result 
            for position in positions: 
                if position != {}: 
                    contracts_check = position 
                    break # exit the for loop 
            # If nothing was found 
            if contracts_check == {}: 
                sell_run_flag = False
            else: 
                if market == 'USD-BTC': 
                    contracts = contracts_check['contracts'] 
                else: 
                    contracts = contracts_check['contracts_no'] 
                balance_available = contracts
                balance_adjust = 0 
                sell_portion = balance_available
        else: # for the other exchanges 
            #print ">> Balance_available pre", balance_available    #DEBUG  
            #print  ">> Balance_adjust pre", balance_adjust     #DEBUG  
            
            # Adjusting according to the limit 
            balance_available = balance_available - Decimal(str(balance_adjust))
            if sell_portion == None: 
                sell_portion = balance_available           

        # Check if we have sold everything 
        if balance_available <= balance_start * Decimal(0.01):
            sell_run_flag = False
        
        # Error strings for exchanges 
        err_1 = 'DUST_TRADE_DISALLOWED_MIN_VALUE_50K_SAT'
        err_2 = 'MIN_TRADE_REQUIREMENT_NOT_MET'
        
        # 2. If something is still required to be sold
        if sell_run_flag: 
            lprint(["Order amount", balance_available, "at price threshold", at_price, "split on", sell_portion])
            remaining_sell_balance = balance_available   
            if exchange == 'bitmex': 
                sale_steps_no = 1       # for the whole position (at least for now) 
            else: 
                sale_steps_no = int(math.ceil(round(Decimal(str(balance_available))/Decimal(str(sell_portion)), 3)))   
            #print ">> Sell amount", balance_available, "remaining_sell_balance", remaining_sell_balance  #DEBUG#
            
            # Selling loop 
            for i in range(1, sale_steps_no + 1):                
                # Check how much should we sell at this step
                if sell_portion > remaining_sell_balance: 
                    sell_q_step = remaining_sell_balance
                else:
                    sell_q_step = sell_portion
                
                # Price update
                price_last_sell = get_last_price(market)
                # Decreasing the price if necessary
                price_to_sell = price_last_sell*(1 - decrease_price_step)
                lprint(["Placing SELL order: Q:", sell_q_step, "@", price_to_sell, "Last market price:", price_last_sell, 
                       "Remaining balance after sale:", round(remaining_sell_balance - sell_q_step, 6)])
                
                # Actually place sell orders if we are not in the simulation mode - re-check
                if simulation != True: 
                    # For bitmex, we will be placing contracts in the other direction (short)
                    if exchange == 'bitmex': 
                        # Balance_available is the number of contracts here. Creating orders depending on the side (long or short) 
                        if market == 'USD-BTC': 
                            price_to_sell = round(price_to_sell, 0)
                        else: 
                            price_to_sell = round(price_to_sell, 20)
                        bitmex_sell_avg_arr.append(price_to_sell) 
                        
                        if short_flag != True: #LONG
                            sell_result = selllimit(exchange, market, sell_q_step, price_to_sell, balance_available) 
                        else: # SHORT   
                            sell_result = buylimit(exchange, market, sell_q_step, price_to_sell, balance_available) 
                        # print "selllimit({}, {}, {}, {}, {})".format(exchange, market, sell_q_step, price_to_sell, balance_available) #DEBUG 
                    else: 
                        sell_result = selllimit(exchange, market, sell_q_step, price_to_sell) 
                    
                    lprint(["-------------------------------------------------------------------- \n>> Sell result:", sell_result, "\n--------------------------------------------------------------------"])  # DEBUG # 
                    
                    if (sell_result == err_1) or (sell_result == err_2):
                        sell_run_flag = False
                        stopmessage = 'err_low'
                    else:
                        # Checking if the sell order was placed
                        try: 
                            if 'uuid' not in sell_result.keys():
                                # Issue with placing order
                                # DEBUG # print "Issue"
                                sell_run_flag = False
                                stopmessage = 'no_idea'
                        except:
                            # DEBUG # print "Issue"
                            sell_run_flag = False
                            stopmessage = 'no_idea'
                
                else: 
                    # If in simulation - calculate profit from virtual sale.  
                    if exchange != 'bitmex': 
                        main_curr_from_sell += float(sell_q_step) * price_to_sell 
                        commission_total += float(sell_q_step)*price_to_sell * comission_rate
                    else: 
                        main_curr_from_sell += contracts_start/price_to_sell
                        commission_total = 0 
                    sell_run_flag = False  
                    
                # Update the db with price_last_sell
                sql_string = "UPDATE jobs SET price_curr={}, selling={} WHERE job_id={}".format(round(price_last_sell, 8), 1, job_id)
                rows = query(sql_string)

                # Decrease remaining balance to sell 
                remaining_sell_balance = remaining_sell_balance - sell_q_step

        # Checking Telegram requests and answering 
        approved_flag = check_cancel_flag()
        if approved_flag == False: 
            # Aborting if asked
            sell_run_flag = False
            stopmessage = 'abort_telegram'
        # Change the flag to sleep on the next cycle
        proceed_w_sleep = True    
        
    # Finishing up
    #print "main_curr_from_sell {}, commission_total {}, contracts_start {}".format (main_curr_from_sell,  commission_total, contracts_start) # DEBUG 
    
    # For bitmex 
    bitmex_sell_avg_arr_np = np.array(bitmex_sell_avg_arr)
    bitmex_sell_avg = bitmex_sell_avg_arr_np.mean()
    
    return stopmessage

################################ Functions - system ############################################
def terminate_w_message(short_text, errtext):
    global logger, handler    
    lprint([short_text])
    send_notification(short_text, errtext)
    logger.close_and_exit()

# Checking if we need to terminate
def check_cancel_flag():
    global market 
    global db, cur, job_id
    
    keep_running = True 
    sql_string = "SELECT abort_flag FROM jobs WHERE job_id = '{}'".format(job_id)
    rows = query(sql_string)

    try: 
        flag_terminate = rows[0][0] # first result 
    except: 
        flag_terminate = 0
    if (flag_terminate == 1): 
        keep_running = False
    return keep_running
 
# Checking if we need to initiate selling from the main or from the mooning cycle 
def check_sell_flag():
    global market 
    global db, cur, job_id
    
    sell_initiate = False 
    sql_string = "SELECT selling FROM jobs WHERE market = '{}'".format(market)
    rows = query(sql_string)

    try: 
        sell_flag = rows[0][0] # first result 
    except: 
        sell_flag = 0
    if (sell_flag == 1): 
        sell_initiate = True
    return sell_initiate
 
def send_notification(subj, text):
    global send_messages, trade_id, comm_method, market
    
    if send_messages:
        if comm_method == 'mail':
            msg = MIMEMultipart()
            msg['From'] = fromaddr
            msg['To'] = toaddr
            msg['Subject'] = trade_id + ': ' + subj
            body = text
            msg.attach(MIMEText(body, 'plain'))
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(fromaddr, email_passw)
            text = msg.as_string()
            server.sendmail(fromaddr, toaddr, text)
            server.quit()  
        else: 
            chat.send(text)
            
def timenow():
    return strftime("%Y-%m-%d %H:%M:%S", localtime())

###################################################################################
############################## Main workflow #########################################
###################################################################################

### Greetings (for logs readability) 
lprint(["###################### ROBOT ###########################"])

if stop_loss: 
    lprint([market, "| Take profit target:", price_target, "| Stop loss:", sl_target, "| Simulation mode:", simulation])
else: 
    lprint([market, "| Take profit target:", price_target, "| Stop loss: disabled (post-profit only) | Simulation mode:", simulation])

if limit_sell_amount > 0: 
    lprint(["Maximum quantity to sell", limit_sell_amount])

### Set up the margin on bitmex 
if exchange == 'bitmex': 
    set_margin = bitmex_leverage(market, bitmex_margin)
    
'''
######## Removed this - will not really be using SL (at least for now), also not applicable for shorts 
# Check if TP is set higher than SL 
if tp < sl: 
    # print "TP {}, SL {}".format(tp, sl) # DEBUG #
    lprint(["Take profit lower than stop loss, r u ok?"])
    exit(0)
''' 

time_hour = time.strftime("%H")     # For periodic updates of 4H candles and stops 
  
# 1. Checking market correctness and URL validity, as well as protecting from fat fingers
try: 
    #ticker_upd = getticker(exchange, market) 
    ticker_upd = coinigy.price(exchange_abbr, market)
    # Ticker could be failing if there is automatic maintenance - then sleep for a while
    if ticker_upd is None: 
        send_notification('Maintenance', market + ' seems to be on an automatic maintenance. Will try every 5 minutes.')
        while ticker_upd is None: 
            lprint(["Market could be on maintenance. Sleeping for 5 minutes."])    
            time.sleep(300) # sleeping for 5 minutes and checking again
            #ticker_upd = getticker(exchange, market) 
            ticker_upd = coinigy.price(exchange_abbr, market)
        
    if ticker_upd == 'INVALID_MARKET': 
        lprint(['Error: Invalid market'])
        logger.close_and_exit()

    else:
        # Fat fingers protection    
        price_check = ticker_upd
        ratio = float(price_target)/float(price_check)
        if (ratio >= 8) or (ratio <= 0.15): 
            lprint(['Error: Double-check prices, are you missing a zero or adding an extra one? The current price is', price_check])
            logger.close_and_exit()

except urllib2.URLError:
    terminate_w_message('Exchange url unavailable')
    logger.close_and_exit()
   
### 2. Checking available balance. If bitmex - checking whether we have a long or a short position 
if simulation != True: 
    balance = getbalance(exchange, currency)
    if balance['Available'] == 0: 
        terminate_w_message('Error: Zero balance', currency + ': zero balance')
        logger.close_and_exit()
    if exchange == 'bitmex': 
        bitmex_no_positions = True 
        positions = getpositions(exchange, market) 
        #print "getpositions({}, {})".format(exchange, market)   #HERE 
        for position in positions: 
            #print position  #HERE
            if position != {}: 
                bitmex_no_positions = False 
                if position['type'] == 'short':     # enabling short flag if the positions are in short 
                    short_flag = True 
                    lprint(['Enabling short_flag'])
                break # exit the for loop 
        if bitmex_no_positions: 
            terminate_w_message('Error: Zero positions on bitmex', 'Error: Zero positions on bitmex')
            logger.close_and_exit()
else: # if simulation 
    if price_target < sl_target: 
        short_flag = True
    else: 
        short_flag = False 
            
### 3. Start the main workflow
run_flag = True 
approved_flag = True
no_sell_orders = False      # default value to suit both simulation and the real run

### 4. Inserting in the sqlite db if started fine ##
sql_string = "INSERT INTO jobs(market, tp, sl, simulation, mooning, selling, price_curr, percent_of, abort_flag, stop_loss, entry_price, mode, tp_p, sl_p, exchange) VALUES ('{}', {}, {}, {}, {},  {},  {},  {},  {}, {}, {}, '{}', {}, {}, '{}')".format(
    market.upper(), price_target, sl_target, int(simulation), int(False), int(False), price_curr, 100, int(False), int(stop_loss), price_entry, simulation_param, tp_p, sl_p, exchange)    
job_id, rows = query_lastrow_id(sql_string)

### 5. Price data for time analysis and strategy. Check what's up with TD analysis data 
start_time = time.time()
td_data_available = True  # default which will be changed to False when needed  
try: 
    bars = td_info.stats(market, exchange_abbr, td_period, 10000, 10, short_flag, market_ref, exchange_abbr_ref)    
    try: 
        if bars == None: 
            td_data_available = False 
    except: 
        for elem in bars['td_setup'][-3:]:      # should have at least 3 bars with filled TD values
            if elem is None: 
                td_data_available = False 
        num_null = bars['open'].isnull().sum()
        if num_null > 0: 
            td_data_available = False 
except: 
    td_data_available = False 
    
print "TD data availability:", td_data_available
# Changing the flip flag to True if td data is not available 
if not td_data_available: 
    price_flip = True
    
### 7. 4H-based stop loss update    
if td_data_available: 
    lprint(["Reconfiguring stop loss level based on TD candles"])
    price_flip_upd, sl_target_upd, sl_upd, sl_p_upd, sl_extreme_upd = stop_reconfigure('now')
    #print ">>> Returned price_flip {}, sl_target_upd {}, sl_upd {}, sl_p_upd {}, sl_extreme_upd {}".format(price_flip_upd, sl_target_upd, sl_upd, sl_p_upd, sl_extreme_upd)  #DEBUG
    if sl_target_upd is not None: 
        sl_target = sl_target_upd
        sl = sl_upd
        sl_p = sl_p_upd    
    if sl_extreme_upd is not None: 
        sl_extreme = sl_extreme_upd
    if price_flip_upd is not None: 
        price_flip = price_flip_upd
        
### 8. Creating new set to store previously executed orders. Will be used to calculate the gains 
orders_start = set()
orders_new = set()

orders_opening = None   #sometimes api fails and ends with an error - so retrying here
while orders_opening is None:
    try:
        orders_opening = getorderhistory(exchange, market)
    except:
         time.sleep(1) 
 
lprint(["Last orders when starting the script"])
if len(orders_opening) < 5: 
    count_max = len(orders_opening)
else: 
    count_max = 5 

for i in range (0, count_max): 
    lprint(['>', orders_opening[i]['OrderUuid']])  
 
for elem in orders_opening: 
    orders_start.add(elem['OrderUuid'])
    #lprint(['>', elem['OrderUuid']]) #DEBUG

# Flags to notify if the prices dropped 
flag_notify_m = True
flag_notify_h = True
dropped_flag = False
    
### 9. Start the main cycle
while run_flag and approved_flag:  
    try:    # try & except is here to raise keyboard cancellation exceptions
        if td_data_available:         # update the stop loss level if due and if we have data
            price_flip_upd, sl_target_upd, sl_upd, sl_p_upd, sl_extreme_upd = stop_reconfigure()
            #print ">>> Returned price_flip {}, sl_target_upd {}, sl_upd {}, sl_p_upd {}, sl_extreme_upd {}".format(price_flip_upd, sl_target_upd, sl_upd, sl_p_upd, sl_extreme_upd) #DEBUG 
            if sl_target_upd is not None: 
                sl_target = sl_target_upd
                sl = sl_upd
                sl_p = sl_p_upd    
                sql_string = "UPDATE jobs SET sl={}, sl_p={} WHERE job_id={}".format(sl_target, sl_p, job_id)   # updating the DB 
                rows = query(sql_string)
            if sl_extreme_upd is not None: 
                sl_extreme = sl_extreme_upd
            if price_flip_upd is not None: 
                price_flip = price_flip_upd
        
        # Get the last price
        price_last = get_last_price(market)
        price_compared = round((float(price_last)/float(price_curr))*100, 2)
        lprint([exchange, market, ": updating price information:", price_last, "|", price_compared, "% of entry price | sl:", sl_target, "sl_extreme", sl_extreme, "| price flip:", price_flip])
        sql_string = "UPDATE jobs SET price_curr={}, percent_of={} WHERE job_id={}".format(round(price_last, 8), price_compared, job_id)
        rows = query(sql_string)
        
        # Commenting for now because the current approach is always time & price based and the approach should not change when the TP is reached 
        ''' 
        ### Running the main conditions check to trigger take profit / stop loss 
        ## Checking TP reached - for longs / shorts 
        if (short_flag and (price_last < price_target)) or ((not short_flag) and (price_last > price_target)):      
            lprint(["Take-profit price reached"])
            send_notification("Mooning", market + ": Reached the initial TP target and mooning now: " + str(price_last))
            status = to_the_moon(price_last)    # mooning for as long as possible 
            if status == 'stop':
                lprint(["Stopped monitoring and finished trades (post-profit)"])
                sleep_timer = 0
                run_flag = False 
                stopped_mode = 'post-profit'    # used in buyback
            elif status == 'abort_telegram': 
                lprint(["Stopped monitoring and finished trades (as requested)"])
                sleep_timer = 0
                run_flag = False 
                stopped_mode = 'telegram' 
        ''' 
                
        ## Pierced through stop loss with flip if stop loss is enabled 
        if stop_loss: 
            # Checking for sequential candle-based signals 
            if (short_flag and price_flip and (price_last >= sl_target)) or ((not short_flag) and price_flip and (price_last <= sl_target)):      
                dropped_flag = True     # changing the flag 
                lprint(["Hitting pre-moon stop loss threshold:", sl_target])
                # Check if we need to sell 
                if not td_data_available: 
                    sale_trigger = ensure_sale(sl_target)   
                else: 
                    sale_trigger = True                
                lprint(["Sale (stop) trigger (pre-moon):", sale_trigger])
                chat.send(market +": exiting based on the time candles (shorter period)")
                 
            # Checking for extreme moves 
            if sl_extreme is not None:
                if (short_flag and (price_last > sl_extreme)) or (not short_flag and (price_last < sl_extreme)): 
                    lprint(["Breached TD extreme stop", price_last])
                    # Check if we need to sell
                    if not td_data_available: 
                        sale_trigger = ensure_sale(price_last)   
                    else: 
                        sale_trigger = True             
                    lprint(["Sale trigger", sale_trigger])
                    chat.send(market +": exiting based on the breach of extreme")
        
            # Checking for nines and breaches of nines, if [-2] (before last) is nine and the price goes beyond the extreme of nine + (or -) contingency
            # bars_4h is a global var in the stop_reconfigure procedure
            if td_data_available: 
                if bars_4h['td_setup'].iloc[-2] == '9': 
                    if short_flag: #shorts 
                        if price_last > bars_4h['high'].iloc[-2]*(1 + var_contingency):
                            sale_trigger = True   
                    else: #longs 
                        if price_last < bars_4h['low'].iloc[-2]*(1 - var_contingency):
                            sale_trigger = True      
                if sale_trigger: 
                    lprint(["Sale trigger based on 9 - 1 rule"])
                    chat.send(market +": exiting based on 9 - 1 rule")
                 
            ### Stop loss triggered 
            if sale_trigger == True:       
                # Stop-loss triggered
                lprint(["Triggering pre-profit stop loss on", price_last])
                send_notification('Sell: SL', exchange + ' : ' + market + ': Triggering pre-moon stop loss at the level of ' + str(price_last))
                status = sell_now(price_last)
                # Handling results
                run_flag, stat_msg = process_stat(status)
                lprint([stat_msg])
                stopped_mode = 'pre-profit'     # used in buyback
                stopped_price = sl_target          
           
           
        # Check if selling now request has been initiated
        sell_init_flag = check_sell_flag()
        if sell_init_flag and approved_flag and run_flag:       
            lprint(["Sale initiated via Telegram @", price_last])
            status = sell_now(price_last)
            # Handling results
            run_flag, stat_msg = process_stat(status)
            lprint([stat_msg])
            stopped_price = price_last  # used in buyback
            stopped_mode = 'telegram' 
        
        # Checking cancellation request and sleeping 
        if run_flag and approved_flag:
            approved_flag = check_cancel_flag()
            if approved_flag == False: 
                lprint(["Shutdown was requested via Telegram"])   
                stopped_mode = 'telegram' 
                sleep_timer = 0
            time.sleep(sleep_timer)
            
    except KeyboardInterrupt:
        lprint(["Shutdown was initiated manually, canceling orders and terminating now"])   
        sql_string = "DELETE FROM jobs WHERE job_id = {}".format(job_id)    # deleting the task from the db 
        rows = query(sql_string)

        # Cancelling orders if not in the simulation mode
        if simulation != True:
            cancel_orders(market)
            time.sleep(10) # wait for cancellations to be processed
            # Information on orders performed if not in a simulation mode
            sell_orders_info()
            sell_orders_outcome()
        logger.close_and_exit()


### 10. Exit point for the main cycle, sell cycle, mooning cycle 
sql_string = "DELETE FROM jobs WHERE job_id = {}".format(job_id)    # deleting the task from the db 
rows = query(sql_string)

# Just a simulation and cancelled by Telegram thing - no virtual sell orders
if simulation == True and approved_flag != True:   
    no_sell_orders = True

### 11. Getting information on performed sell orders and displaying / recording the outcomes
sell_orders_info()
sell_orders_outcome()

# Then start monitoring for buyback (both post-moon and SL)  'pre-profit'  /  'post-profit'
''' 
# Uncomment if you would like to restrict buybacks, code should also be revised from long / short perspective 
# Checking losses to stop buyback in case of 2 consecutive losses incurred  
sql_string = "SELECT id FROM losses WHERE market = '{}'".format(market)
rows = query(sql_string)
try: 
    loss_id = int(rows[0][0]) # if one loss already have been incurred - no bback 
except: 
    loss_id = None
''' 
if (stopped_mode == 'pre-profit') and (cancel_buyback == False): 
    if short_flag != True: # LONGS  
        bb_price = price_exit * 0.9975  # Using value of price_exit (actual sell price) minus 0.25% (for non-TD-based price action) 
    else: 
        bb_price = price_exit * 1.0025 # SHORTS  
    if td_data_available:  
        lprint(["Buyback will be based on TD candles"])
    else: 
        lprint(["Setting buyback price as actual sell +/- 0.25%. Price_exit:", price_exit, "bb_price", bb_price])
    
    '''
    # Uncomment if you would like to restrict buybacks, code should also be revised from long / short perspective 
    # Checking losses to stop buyback in case of 2 consecutive losses incurred  
    if loss_id is not None: 
        chat.send(market +": stopped buyback after two consecutive losses")
        sql_string = "DELETE FROM losses WHERE id = {}".format(loss_id)
        rows = query(sql_string)
        logger.close_and_exit()
        exit(0) 
    else: 
        # Inserting into losses table if this is the first occurence 
        sql_string = "INSERT INTO losses(market) VALUES ('{}')".format(market)
        rows = query(sql_string)    
    ''' 
elif stopped_mode == 'post-profit': 
    # Thresholds for post-profit fallback for BTC or ALTS
    if market == 'USDT-BTC': 
        if short_flag != True: # LONGS  
            bb_price = price_exit  * 1.005 # Using fixed value of +0.5% from stop; however, does not refer to price value when using TD analysis 
        else: 
            bb_price = price_exit  * 0.995
    else: 
        if short_flag != True: # LONGS  
            bb_price = price_exit  * 1.01  
        else: 
            bb_price = price_exit  * 0.99
     
    if td_data_available:  
        lprint(["Buyback will be based on TD candles"])
    else: 
        lprint(["Setting buyback price as actual +/- 1%. Stopped_price:", stopped_price, "bb_price", bb_price])
    ''' 
    # Uncomment if you would like to restrict buybacks 
    # If it was a loser - delete the info in DB and continue with BBack 
    if loss_id is not None: 
        sql_string = "DELETE FROM losses WHERE id = {}".format(loss_id)
        rows = query(sql_string)
    ''' 
else: 
    # If just called for stop from Telegram 
    lprint(["Sold through telegram"])   
    if price_exit is None: 
        bb_price = price_last
    else: 
        bb_price = price_exit

### 12. Buying back based on 4H action or alternative price action    
try: 

    # Starting buyback except for the cases when the task was aborted through telegram 
    if stopped_mode != 'telegram':      
        lprint(["Buyback monitoring started:", stopped_mode, "| TD data availability:", td_data_available])   
        
        if exchange == 'bitmex': 
            buy_trade_price = (main_curr_from_sell)/bitmex_margin
        else: 
            buy_trade_price = float(balance_start) * bb_price * (1 - comission_rate)    # commission depending on the exchange. If we do not have TD data
        
        # Inserting into buyback information table 
        sql_string = "INSERT INTO bback(market, bb_price, curr_price, trade_price, exchange) VALUES ('{}', {}, {}, {}, '{}')".format(market, bb_price, bb_price, buy_trade_price, exchange)
        bb_id, rows = query_lastrow_id(sql_string)      
        
        time_bb_initiated = time.time()     # getting a snapshot of time for buyback so that we wait for at least an hour before starting buyback 
        bb_flag, direction = buy_back(bb_price)      # runs until a result is returned with a confirmation and a direction which defines the next step  
        
        # If we have reached the target to initiate a buyback and there was no cancellation through Telegram
        if bb_flag: 
            send_notification('Buyback', 'Buy back initiated for ' + market + ' on ' + exchange + '. Direction: ' + direction)  
            
            # Launching workflow to buy and resume the task with same parameters
            # Insert a record in the db: workflow(wf_id INTEGER PRIMARY KEY, tp FLOAT, sl FLOAT, sell_portion FLOAT)
            
            if direction == 'up': #LONGS 
                sl_price = bb_price * (1 - diff_threshold)  # depending on the strategy 
                tp_price = bb_price * tp
            elif direction == 'down': 
                sl_price = bb_price * (1 + diff_threshold)  # depending on the strategy 
                tp_price = bb_price / tp

            #print "bb_price {}, tp {}, diff_threshold {}, tp_price {}, sl_price {}".format(bb_price, tp, diff_threshold, tp_price, sl_price) # DEBUG 
                
            sql_string = "INSERT INTO workflow(tp, sl, sell_portion, run_mode, price_entry, exchange) VALUES ({}, {}, {}, '{}', {}, '{}')".format(tp_price, sl_price, 0, simulation_param, float(bb_price), exchange_abbr)
            wf_id, rows = query_lastrow_id(sql_string)       

            if wf_id is not None: 
                buy_market = '{0}-{1}'.format(trade, currency)
                sql_string = "UPDATE workflow SET market = '{}', trade = '{}', currency = '{}', exchange = '{}' WHERE wf_id = {}".format(market, trade, currency, exchange_abbr, wf_id) 
                job_id, rows = query_lastrow_id(sql_string)
                
            # Buy depending on the platform. We will buy @ market price now, and the price entry price is already in the DB
            if td_data_available: 
                mode_buy = 'now' 
            else: 
                mode_buy = 'reg' 

            logger.close() # closing logs 
            
            sql_string = "DELETE FROM bback WHERE id = {}".format(bb_id)  # deleting buyback from the table 
            rows = query(sql_string)
            
            # Run a smart buy task now when we have a buyback confirmation 
            if direction == 'up': #LONGS 
                python_call = 'python smart_buy.py ' + ' '.join([mode_buy, exchange_abbr, trade + '-' + currency, str(buy_trade_price)])
            elif direction == 'down': #SHORTS - need to change plus to minus
                python_call = 'python smart_buy.py ' + ' '.join([mode_buy, exchange_abbr, trade + '-' + currency, str(-buy_trade_price)])
            print '>>>' + python_call
            p = subprocess.Popen(python_call, shell=True, stderr=subprocess.PIPE)
            while True:
                out = p.stderr.read(1)
                if out == '' and p.poll() != None:
                    break
                if out != '':
                    sys.stdout.write(out)
                    sys.stdout.flush()
        
        # If a buyback cancellation was requested 
        else: 
            send_notification('Buyback', 'Buy back cancelled as requested for ' + market + ' on ' + exchange) 
            # Delete buyback from the DB 
            sql_string = "DELETE FROM bback WHERE id = {}".format(bb_id)
            rows = query(sql_string)
    
    # If telegram stop - finalise and exit 
    else:  
        logger.close_and_exit()
    
except KeyboardInterrupt:
    print "Buyback cancelled or the task was finished"  
     # Delete buyback from the DB 
    sql_string = "DELETE FROM bback WHERE id = {}".format(bb_id)
    rows = query(sql_string)
    try: 
        logger.close_and_exit()
    except: 
        print 'Logs already closed' 
